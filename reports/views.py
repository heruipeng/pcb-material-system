from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django.utils import timezone
from django.db.models import Count
from datetime import datetime, timedelta

from .models import Report, ReportCategory, ReportInstance, Dashboard, ScheduledReport
from .serializers import (
    ReportSerializer, ReportCategorySerializer, ReportInstanceSerializer,
    DashboardSerializer, ScheduledReportSerializer
)
from core.utils import log_operation


class ReportCategoryViewSet(viewsets.ModelViewSet):
    """报表分类管理"""
    queryset = ReportCategory.objects.filter(is_active=True)
    serializer_class = ReportCategorySerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name', 'code']
    ordering_fields = ['sort_order', 'id']
    ordering = ['sort_order', 'id']


class ReportViewSet(viewsets.ModelViewSet):
    """报表管理 - 完整 CRUD + 生成 Excel"""
    queryset = Report.objects.select_related('category', 'created_by').filter(is_active=True)
    serializer_class = ReportSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['report_type', 'category', 'is_system']
    search_fields = ['name', 'code', 'description']
    ordering_fields = ['created_at', 'updated_at', 'name', 'sort_order']
    ordering = ['sort_order', 'id']

    @action(detail=True, methods=['post'])
    def generate(self, request, pk=None):
        """生成报表 - 使用报表定义的查询条件或用户提交的过滤参数"""
        import os
        from django.conf import settings
        from django.db.models import Q
        from openpyxl import Workbook

        report = self.get_object()
        
        # 从请求或报表定义中获取查询条件
        date_from = request.data.get('date_from', report.query_params.get('date_from', ''))
        date_to = request.data.get('date_to', report.query_params.get('date_to', ''))
        factory_id = request.data.get('factory', report.query_params.get('factory', ''))
        status_filter = request.data.get('status', report.query_params.get('status', ''))
        process_type = request.data.get('process_type', report.query_params.get('process_type', ''))

        # 根据查询条件获取实际数据
        from materials.models import Material
        queryset = Material.objects.select_related('factory', 'maker', 'creator', 'category').all()
        
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        if factory_id:
            queryset = queryset.filter(factory_id=factory_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if process_type:
            queryset = queryset.filter(process_type=process_type)
        
        materials = list(queryset[:1000])  # 限制导出上限
        row_count = len(materials)

        # 生成 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.title = report.name
        headers = ['序号', '料号', '流水号', '工厂', '状态', '工具类型', '版本', '备注']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        for i, m in enumerate(materials):
            ws.cell(row=i+2, column=1, value=i+1)
            ws.cell(row=i+2, column=2, value=m.material_no)
            ws.cell(row=i+2, column=3, value=m.serial_no)
            ws.cell(row=i+2, column=4, value=m.factory.name if m.factory else '')
            ws.cell(row=i+2, column=5, value=m.get_status_display())
            ws.cell(row=i+2, column=6, value=m.get_process_type_display())
            ws.cell(row=i+2, column=7, value=m.version_code)
            ws.cell(row=i+2, column=8, value=m.remark or '')

        # 保存到 media 目录
        media_root = settings.MEDIA_ROOT
        report_dir = os.path.join(media_root, 'reports')
        os.makedirs(report_dir, exist_ok=True)
        filename = f"{report.code}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(report_dir, filename)
        wb.save(filepath)

        # 创建报表实例
        instance = ReportInstance.objects.create(
            report=report,
            name=f"{report.name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}",
            status='completed',
            generated_by=request.user if request.user.is_authenticated else None,
            row_count=row_count,
            query_params={
                'date_from': date_from,
                'date_to': date_to,
                'factory': factory_id,
                'status': status_filter,
                'process_type': process_type,
            },
            date_from=datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None,
            date_to=datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None,
            file=f'reports/{filename}',
            file_format='xlsx',
            file_size=os.path.getsize(filepath),
            generated_at=timezone.now(),
            completed_at=timezone.now(),
        )

        log_operation(request, 'create', 'reports', 'ReportInstance', instance.id,
                     f'生成报表 {report.name}')

        return Response({
            'success': True,
            'instance_id': instance.id,
            'status': 'completed',
            'row_count': row_count,
            'file_url': instance.file.url if instance.file else None,
            'file_name': filename,
        })

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """报表统计"""
        today = timezone.now().date()
        week_start = today - timedelta(days=today.weekday())

        return Response({
            'total_count': Report.objects.count(),
            'active_count': Report.objects.filter(is_active=True).count(),
            'today_count': ReportInstance.objects.filter(generated_at__date=today).count(),
            'week_count': ReportInstance.objects.filter(generated_at__date__gte=week_start).count(),
            'by_type': list(Report.objects.values('report_type').annotate(c=Count('id'))),
        })

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """下载报表最新文件"""
        report = self.get_object()
        instance = report.instances.filter(status='completed').order_by('-generated_at').first()
        if instance and instance.file:
            return Response({'file_url': instance.file.url, 'file_name': instance.file.name})
        return Response({'error': '没有可下载的文件'}, status=status.HTTP_404_NOT_FOUND)

class ReportInstanceViewSet(viewsets.ReadOnlyModelViewSet):
    """报表实例 - 只读"""
    queryset = ReportInstance.objects.select_related('report', 'generated_by').all()
    serializer_class = ReportInstanceSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['report', 'status']
    search_fields = ['name']
    ordering_fields = ['generated_at', 'row_count', 'file_size']
    ordering = ['-generated_at']


class DashboardViewSet(viewsets.ModelViewSet):
    """仪表盘管理"""
    queryset = Dashboard.objects.all()
    serializer_class = DashboardSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['created_at']
    ordering = ['-created_at']


class ScheduledReportViewSet(viewsets.ModelViewSet):
    """定时报表管理"""
    queryset = ScheduledReport.objects.select_related('report').all()
    serializer_class = ScheduledReportSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['report', 'is_active', 'frequency']
    search_fields = ['name']
    ordering_fields = ['created_at']
    ordering = ['-created_at']
